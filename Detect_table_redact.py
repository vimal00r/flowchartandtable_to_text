import fitz
import pandas as pd
import numpy as np
import glob
import requests
import xlsxwriter
import pdfplumber
import openpyxl
import re
import os
import xlsxwriter
import PyPDF2

from bs4 import BeautifulSoup
from img2table.document import PDF
from img2table.ocr import TesseractOCR
from IPython.display import display_html
from bs4 import BeautifulSoup
from fitz.fitz import TEXT_ALIGN_LEFT


def get_tables(text, name):
    soup = BeautifulSoup(text, 'html.parser')  # Specify the parser explicitly
    tables = soup.find_all('table')  # Find all tables in the HTML
    for table_index, table in enumerate(tables):
        header = []
        rows = []
        for i, row in enumerate(table.find_all('tr')):
            cells = [cell.text.strip().replace('\n', '') for cell in row.find_all(['th', 'td'])]
            if i == 0:
                header = cells
            else:
                rows.append(cells)
        # Create a separate XLSX file for each table
        workbook = xlsxwriter.Workbook(name)
        worksheet = workbook.add_worksheet()
        worksheet.write_row(0, 0, header)
        for i, row in enumerate(rows):
            worksheet.write_row(i + 1, 0, row)
        workbook.close()


def get_all_pages_text(pdf_path):
    arr=[]
    def not_within_bboxes(obj):
        def obj_in_bbox(_bbox):
            v_mid = (obj["top"] + obj["bottom"]) / 2
            h_mid = (obj["x0"] + obj["x1"]) / 2
            x0, top, x1, bottom = _bbox
            return (h_mid >= x0) and (h_mid < x1) and (v_mid >= top) and (v_mid < bottom)
        return not any(obj_in_bbox(__bbox) for __bbox in bboxes)
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            try:
                bboxes = [
                    table.bbox
                    for table in page.find_tables(
                        table_settings={"vertical_strategy": "explicit",
                                        "horizontal_strategy": "explicit",
                                        "explicit_vertical_lines": page.curves + page.edges,
                                        "explicit_horizontal_lines": page.curves + page.edges,
                                        }
                    )
                ]
            except Exception as e:
                print(f"An error occurred: Image Appeared")
                continue
            arr.append(page.filter(not_within_bboxes).extract_text())
    return arr


def get_all_content(pdf_path):
    ocr=TesseractOCR(lang='eng')
    pdf=PDF(src=pdf_path)
    extracted_tables = pdf.extract_tables(ocr=ocr,
                                      implicit_rows=False,
                                      borderless_tables=False,
                                      min_confidence=50)
    #Saving html table to xlsx format
    for page, tables in extracted_tables.items():
        for idx, table in enumerate(tables):
            z=table.html_repr()
            get_tables(z,f'./Table_to_text/tables/Page-{page+1} Table-{idx+1}.xlsx')



# Function to convert XLSX to TXT
def xlsx_to_txt(xlsx_file, txt_file):
    try:
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active
        with open(txt_file, 'w', encoding='utf-8') as txt_file:
            # Iterate through rows in the sheet and write cell values to the TXT file
            for row in sheet.iter_rows():
                row_data = [str(cell.value) if cell.value is not None else "" for cell in row]
                row_text = '\t'.join(row_data)  # Separate cell values by tabs
                txt_file.write(row_text + '\n')  # Write the row to the TXT file
    except Exception as e:
        print(f"Error: {e}")



def pdf_redact(doc,page_num,rectangle,text):
    doc[page_num].add_redact_annot(rectangle)
    doc[page_num].apply_redactions()
    tw = fitz.TextWriter(doc[page_num].rect)
    tw.fill_textbox(rectangle,text)
    tw.write_text(doc[page_num])



def read(txt_file_path):
    with open(txt_file_path, "r",encoding='utf-8') as file:
            file_contents = file.read()
    return file_contents


folder_path = './Table_to_text/tables/'

table_id_dict = {}

def table_to_text(pdf_path):
        get_all_content(pdf_path)

        for filename in os.listdir(folder_path):
            if filename.endswith('.xlsx'):
                xlsx_file = os.path.join(folder_path, filename)
                txt_file = os.path.join('./Table_to_text/tables_data/', filename.replace('.xlsx', '.txt'))
                xlsx_to_txt(xlsx_file, txt_file)
                # with open(txt_file, "rb") as file:
                #     file_contents = file.read()
                # table_id_dict[filename.split(".")[0]] = file_contents

        doc=fitz.open(pdf_path)
        with pdfplumber.open(pdf_path) as pdf:
            for num,page in enumerate(pdf.pages):
                k=page.find_tables(table_settings={"vertical_strategy": "explicit",
                                        "horizontal_strategy": "explicit",
                                        "explicit_vertical_lines": page.curves + page.edges,
                                        "explicit_horizontal_lines": page.curves + page.edges,
                                        })
                for table,z in enumerate(k):
                    # pdf_redact(doc,num,fitz.Rect(z.bbox),f"Page-{num+1} Table-{table+1}")
                    pdf_redact(doc,num,fitz.Rect(z.bbox),str(read("./Table_to_text/tables_data/"+f"Page-{num+1} Table-{table+1}.txt")))
                  
        doc.save('./'+pdf_path,deflate=True)
        doc.close()
        # return table_id_dict

table_to_text("OBC_(Onboard_Charger)_Project_Management_I-00347-01-01_SysTS01 (1).pdf")




# pdf_writer = PyPDF2.PdfWriter()
# def extract_text(pdf_file):
#     with pdfplumber.open(pdf_file) as pdf:
#         text1 = ""
#         for page in pdf.pages:
#             text1 += page.extract_text()

#         # Get all files in the directory
#         files = os.listdir('./Table_to_text/tables_data/')
#         for file in files:
#             # Construct the pattern from the filename
#             pattern = file.replace('.txt', '') 
#             filename = './Table_to_text/tables_data/' + file
#             # pdf_filename = './Supertext_output.pdf'
#             with open(filename, "r",encoding='utf-8') as file:
#                 file_text = file.read()
#                 # with open(pdf_filename, "w") as pdffile:
#                 #     pdf_writer.add_page(file_text)
#             text1 = text1.replace(pattern, file_text)

#         # Save text1 to a super text file
#         super_text_filename = "./Table_to_text/super_text.txt"
#         with open(super_text_filename, 'w') as super_text_file:
#             super_text_file.write(text1)

#         return text1

# extract_text("./table_to_text_output.pdf")